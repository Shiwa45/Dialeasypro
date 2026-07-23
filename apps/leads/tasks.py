"""
TeleCRM Backend — apps/leads/tasks.py

Celery tasks for the leads app.

process_lead_import     : Main import pipeline — reads CSV/Excel, validates,
                          deduplicates, bulk-inserts leads, generates error report.
recalculate_lead_scores : Periodic scoring task (runs for all tenants).
send_followup_reminder  : Sends reminder for an upcoming follow-up.
cleanup_old_leads       : Archives/deletes leads past data retention period.
"""
import logging

from celery import shared_task
from django.conf import settings
from django.utils import timezone

from apps.core.tasks import PublicSchemaTask, TenantAwareTask

logger = logging.getLogger(__name__)


# ============================================================
# Lead Import Pipeline
# ============================================================

@shared_task(base=TenantAwareTask, bind=True, max_retries=2, time_limit=1800)
def process_lead_import(self, schema_name: str, import_job_id: str):
    """
    Main lead import pipeline.
    Reads file from storage, validates rows, deduplicates, bulk-inserts.

    Progress is tracked via LeadImportJob model fields (polled by frontend).
    Celery task ID stored in LeadImportJob.celery_task_id.
    """
    import csv
    import io
    from apps.leads.models import Lead, LeadActivity, LeadImportJob
    from apps.core.constants import LeadSource
    from apps.core.utils import normalize_indian_phone

    logger.info(f"[Import] Starting import job {import_job_id} in {schema_name}")

    try:
        job = LeadImportJob.objects.get(pk=import_job_id)
    except LeadImportJob.DoesNotExist:
        logger.error(f"[Import] Job {import_job_id} not found in {schema_name}")
        return

    job.status = "processing"
    job.save(update_fields=["status"])

    # --- Read file ------------------------------------------
    try:
        file_content = job.file.read()
        filename = job.original_filename.lower()
    except Exception as exc:
        job.status = "failed"
        job.row_errors = [{"row": 0, "error": f"Could not read file: {exc}"}]
        job.save(update_fields=["status", "row_errors"])
        return

    # --- Parse into rows ------------------------------------
    rows = _parse_import_file(file_content, filename)
    if rows is None:
        job.status = "failed"
        job.row_errors = [{"row": 0, "error": "Unsupported file format. Use CSV or XLSX."}]
        job.save(update_fields=["status", "row_errors"])
        return

    job.total_rows = len(rows)
    job.save(update_fields=["total_rows"])

    # --- Mapping of column names → Lead fields ---------------
    column_mapping = job.column_mapping or {}
    # Defaults if no mapping configured
    default_mapping = {
        "name": ["name", "full name", "contact name", "customer name", "lead name"],
        "phone": ["phone", "mobile", "contact", "phone number", "mobile number", "contact no"],
        "email": ["email", "email address", "mail"],
        "city": ["city", "location", "town"],
        "state": ["state"],
        "requirement": ["requirement", "query", "message", "remarks", "comments"],
        "budget": ["budget", "amount"],
    }

    # --- Process rows ----------------------------------------
    successful, failed, duplicates = 0, 0, 0
    row_errors = []
    leads_to_create = []

    # Build existing phone set for duplicate detection
    existing_phones = set(
        Lead.objects.filter(is_deleted=False).values_list("phone", flat=True)
    )

    for row_idx, raw_row in enumerate(rows, start=2):  # start=2 (row 1 is header)
        try:
            # Apply column mapping
            row = _apply_mapping(raw_row, column_mapping, default_mapping)

            # Validate required fields
            if not row.get("name", "").strip():
                raise ValueError("Name is required")
            if not row.get("phone", "").strip():
                raise ValueError("Phone is required")

            # Normalize phone
            phone = normalize_indian_phone(row["phone"])
            if not phone:
                raise ValueError(f"Invalid phone number: {row['phone']}")

            # Duplicate handling
            if phone in existing_phones:
                if job.duplicate_action == "skip":
                    duplicates += 1
                    job.processed_rows += 1
                    job.save(update_fields=["processed_rows"])
                    continue
                elif job.duplicate_action == "update":
                    # Update existing lead
                    try:
                        existing = Lead.objects.get(phone=phone, is_deleted=False)
                        _update_lead_from_row(existing, row, job)
                        successful += 1
                        duplicates += 1
                    except Lead.DoesNotExist:
                        pass
                    job.processed_rows += 1
                    job.save(update_fields=["processed_rows"])
                    continue
                # else: create_new — allow duplicate

            # Build lead kwargs
            lead_kwargs = {
                "name": row["name"].strip(),
                "phone": phone,
                "email": row.get("email", "").strip().lower(),
                "city": row.get("city", "").strip(),
                "state": row.get("state", "").strip(),
                "requirement": row.get("requirement", "").strip(),
                "source": job.default_source or LeadSource.MANUAL,
                "assigned_to": job.default_assigned_to,
                "import_job": job,
            }
            # Budget parsing
            budget_str = row.get("budget", "").strip().replace(",", "").replace("₹", "")
            if budget_str:
                try:
                    lead_kwargs["budget"] = float(budget_str)
                except ValueError:
                    pass

            leads_to_create.append(Lead(**lead_kwargs))
            existing_phones.add(phone)

        except Exception as exc:
            failed += 1
            row_errors.append({
                "row": row_idx,
                "data": str(raw_row)[:200],
                "error": str(exc),
            })

        job.processed_rows = row_idx - 1
        # Save progress every 100 rows
        if (row_idx - 1) % 100 == 0:
            job.save(update_fields=["processed_rows"])

    # --- Plan capacity ----------------------------------------
    # Truncate to what the plan still allows rather than failing the whole job:
    # importing 900 of 1000 rows is far more useful than importing none.
    from apps.core.quotas import note_leads_created, remaining_lead_allowance

    allowance = remaining_lead_allowance()
    if allowance is not None and len(leads_to_create) > allowance:
        skipped = len(leads_to_create) - allowance
        leads_to_create = leads_to_create[:allowance]
        failed += skipped
        row_errors.append({
            "row": "-",
            "data": "",
            "error": (
                f"Plan lead limit reached — {skipped} row(s) not imported. "
                f"Upgrade your plan to import more."
            ),
        })
        logger.warning(f"[Import] job={job.pk} truncated {skipped} rows (plan limit)")

    # --- Bulk create -----------------------------------------
    if leads_to_create:
        try:
            chunk_size = 500
            for i in range(0, len(leads_to_create), chunk_size):
                chunk = leads_to_create[i:i + chunk_size]
                Lead.objects.bulk_create(chunk, ignore_conflicts=True)
                # Log activity for created leads
                activities = [
                    LeadActivity(
                        lead=lead,
                        activity_type="imported",
                        description=f"Imported from {job.original_filename}",
                    )
                    for lead in chunk
                    if lead.pk  # pk is set after bulk_create
                ]
                if activities:
                    LeadActivity.objects.bulk_create(activities)
            successful += len(leads_to_create)
            note_leads_created(len(leads_to_create))
        except Exception as exc:
            logger.error(f"[Import] Bulk create failed: {exc}", exc_info=True)
            failed += len(leads_to_create)
            row_errors.append({"row": "bulk", "error": str(exc)})

    # --- Finalize -------------------------------------------
    job.successful_rows = successful
    job.failed_rows = failed
    job.duplicate_rows = duplicates
    job.processed_rows = job.total_rows
    job.row_errors = row_errors[:500]  # Cap error list size
    job.mark_completed()

    logger.info(
        f"[Import] Job {import_job_id} done: "
        f"{successful} ok / {failed} failed / {duplicates} dupes"
    )
    return {"successful": successful, "failed": failed, "duplicates": duplicates}


def _parse_import_file(file_content: bytes, filename: str) -> list | None:
    """
    Universal sheet & data file parser.
    Supports CSV, XLSX, XLS, XLSM, XLSB, ODS, TSV, TXT files regardless of case or extension.
    Handles encodings, delimiters, float phone numbers (e.g. 9876543210.0 -> 9876543210),
    and dates automatically.
    """
    import io
    import csv
    import logging

    logger = logging.getLogger(__name__)
    if not file_content:
        return None

    fname_lower = (filename or "").lower()
    is_excel = file_content.startswith(b'PK\x03\x04') or file_content.startswith(b'\xd0\xcf\x11\xe0') or fname_lower.endswith((".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"))

    # ============================================================
    # 1. EXCEL PARSING (.xlsx, .xls, .xlsm, .ods)
    # ============================================================
    if is_excel:
        # Strategy A: Openpyxl (standard mode)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    rows.append(row)
            if rows and len(rows) >= 2:
                headers = [str(c or "").strip() for c in rows[0]]
                parsed_rows = []
                for row_cells in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(row_cells):
                        if i < len(headers) and headers[i]:
                            if val is None:
                                v_str = ""
                            elif isinstance(val, float) and val.is_integer():
                                v_str = str(int(val))
                            else:
                                v_str = str(val).strip()
                                if v_str.endswith(".0") and v_str[:-2].isdigit():
                                    v_str = v_str[:-2]
                            row_dict[headers[i]] = v_str
                    if any(v for v in row_dict.values()):
                        parsed_rows.append(row_dict)
                if parsed_rows:
                    return parsed_rows
        except Exception as exc:
            logger.info(f"[Import] openpyxl standard load failed: {exc}")

        # Strategy B: Openpyxl (read_only mode)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            iterator = ws.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if header_row:
                headers = [str(c or "").strip() for c in header_row]
                parsed_rows = []
                for row in iterator:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and headers[i]:
                            if val is None:
                                v_str = ""
                            elif isinstance(val, float) and val.is_integer():
                                v_str = str(int(val))
                            else:
                                v_str = str(val).strip()
                                if v_str.endswith(".0") and v_str[:-2].isdigit():
                                    v_str = v_str[:-2]
                            row_dict[headers[i]] = v_str
                    if any(v for v in row_dict.values()):
                        parsed_rows.append(row_dict)
                if parsed_rows:
                    return parsed_rows
        except Exception as exc:
            logger.info(f"[Import] openpyxl read_only load failed: {exc}")

        # Strategy C: Pandas read_excel
        try:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(file_content), dtype=str)
            df = df.dropna(how="all")
            df.columns = [str(col).strip() for col in df.columns]
            parsed_rows = []
            for _, row in df.iterrows():
                row_dict = {}
                for col in df.columns:
                    val = row[col]
                    if pd.isna(val) or val is None or str(val).strip().lower() in ("nan", "none", "null", "nat"):
                        val_str = ""
                    else:
                        val_str = str(val).strip()
                        if val_str.endswith(".0") and val_str[:-2].replace("-", "").isdigit():
                            val_str = val_str[:-2]
                    row_dict[col] = val_str
                if any(v for v in row_dict.values()):
                    parsed_rows.append(row_dict)
            if parsed_rows:
                return parsed_rows
        except Exception as exc:
            logger.info(f"[Import] pandas read_excel failed: {exc}")

    # ============================================================
    # 2. CSV / TEXT PARSING (.csv, .tsv, .txt, or text bytes)
    # ============================================================
    for encoding in ["utf-8-sig", "utf-8", "cp1252", "latin-1", "iso-8859-1", "gbk"]:
        try:
            text = file_content.decode(encoding)
        except Exception:
            continue

        # Skip binary garbage if accidentally passed here
        if "\0" in text[:500]:
            continue

        sample = text[:4096]
        # Detect delimiter
        delimiter = ","
        if ";" in sample and sample.count(";") > sample.count(","):
            delimiter = ";"
        elif "\t" in sample and sample.count("\t") > sample.count(","):
            delimiter = "\t"
        elif "|" in sample and sample.count("|") > sample.count(","):
            delimiter = "|"

        try:
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            rows = []
            for row in reader:
                cleaned_row = {}
                for k, v in row.items():
                    if k is not None:
                        k_str = str(k).strip()
                        v_str = str(v or "").strip()
                        if v_str.endswith(".0") and v_str[:-2].isdigit():
                            v_str = v_str[:-2]
                        cleaned_row[k_str] = v_str
                if any(v for v in cleaned_row.values()):
                    rows.append(cleaned_row)
            if rows:
                return rows
        except Exception as exc:
            logger.info(f"[Import] CSV decode/parse attempt ({encoding}, {delimiter}) failed: {exc}")

    # Strategy D: Pandas read_csv fallback for complex CSVs
    try:
        import pandas as pd
        for encoding in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
            for sep in [",", ";", "\t", "|"]:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), sep=sep, encoding=encoding, dtype=str, on_bad_lines="skip")
                    df = df.dropna(how="all")
                    if not df.empty:
                        df.columns = [str(c).strip() for c in df.columns]
                        parsed_rows = []
                        for _, row in df.iterrows():
                            r_dict = {col: ("" if pd.isna(row[col]) else str(row[col]).strip()) for col in df.columns}
                            if any(v for v in r_dict.values()):
                                parsed_rows.append(r_dict)
                        if parsed_rows:
                            return parsed_rows
                except Exception:
                    continue
    except Exception as exc:
        logger.info(f"[Import] pandas read_csv fallback failed: {exc}")

    return None


def _apply_mapping(raw_row: dict, column_mapping: dict, default_mapping: dict) -> dict:
    """Map raw CSV column names to Lead field names."""
    result = {}

    if column_mapping:
        # User-configured explicit mapping
        for lead_field, csv_col in column_mapping.items():
            result[lead_field] = raw_row.get(csv_col, "")
    else:
        # Auto-detect by normalizing column names
        normalized_row = {k.lower().strip(): v for k, v in raw_row.items()}
        for lead_field, aliases in default_mapping.items():
            for alias in aliases:
                if alias in normalized_row:
                    result[lead_field] = normalized_row[alias]
                    break

    return result


def _update_lead_from_row(lead: "Lead", row: dict, job: "LeadImportJob"):
    """Update an existing lead from import row data (duplicate update mode)."""
    update_fields = []
    if row.get("email") and not lead.email:
        lead.email = row["email"].strip().lower()
        update_fields.append("email")
    if row.get("city") and not lead.city:
        lead.city = row["city"].strip()
        update_fields.append("city")
    if row.get("requirement") and not lead.requirement:
        lead.requirement = row["requirement"].strip()
        update_fields.append("requirement")
    if update_fields:
        lead.save(update_fields=update_fields)


# ============================================================
# Lead Scoring
# ============================================================

@shared_task(base=TenantAwareTask, bind=True)
def recalculate_lead_scores(self, schema_name: str):
    """
    Recalculate lead quality scores for all active leads.
    Scoring factors:
    - Has phone: +20
    - Has email: +10
    - Has budget set: +15
    - Has requirement: +10
    - Source quality: +0-20 (IndiaMART/Meta > manual > others)
    - Follow-up completed: +10 per completed (max 20)
    - Response rate: +0-15
    - Days old penalty: -1 per 7 days (max -30)
    """
    from apps.leads.models import Lead, FollowUp
    from apps.core.constants import LeadSource, LeadStatus

    EXCLUDED_STATUSES = [LeadStatus.WON, LeadStatus.LOST, LeadStatus.NOT_INTERESTED]
    SOURCE_SCORES = {
        LeadSource.INDIAMART: 20,
        LeadSource.META_FACEBOOK: 18,
        LeadSource.GOOGLE_ADS: 18,
        LeadSource.WEBSITE: 15,
        LeadSource.REFERRAL: 20,
        LeadSource.MANUAL: 10,
    }

    active_leads = Lead.objects.filter(
        is_deleted=False
    ).exclude(status__in=EXCLUDED_STATUSES)

    update_list = []
    for lead in active_leads.iterator(chunk_size=200):
        score = 0
        if lead.phone:
            score += 20
        if lead.email:
            score += 10
        if lead.budget:
            score += 15
        if lead.requirement:
            score += 10
        score += SOURCE_SCORES.get(lead.source, 5)
        completed_fus = FollowUp.objects.filter(lead=lead, is_completed=True).count()
        score += min(20, completed_fus * 10)
        days_old = (timezone.now() - lead.created_at).days
        score -= min(30, (days_old // 7))

        lead.score = max(0, min(100, score))
        update_list.append(lead)

        if len(update_list) >= 500:
            Lead.objects.bulk_update(update_list, ["score"])
            update_list = []

    if update_list:
        Lead.objects.bulk_update(update_list, ["score"])

    logger.info(f"[Task] Lead scores updated for {schema_name}")


# ============================================================
# Follow-up Reminder
# ============================================================

@shared_task(base=TenantAwareTask, bind=True, max_retries=2)
def send_followup_reminders_for_tenant(self, schema_name: str):
    """
    Find all follow-ups due in the next 30 minutes and send reminders.
    Called from apps.authentication.tasks.dispatch_followup_reminders.
    """
    from datetime import timedelta
    from apps.leads.models import FollowUp
    from apps.core.consumers import send_agent_notification

    now = timezone.now()
    window_end = now + timedelta(minutes=30)

    due_followups = FollowUp.objects.filter(
        is_completed=False,
        reminder_sent=False,
        scheduled_at__gte=now,
        scheduled_at__lte=window_end,
    ).select_related("lead", "assigned_to")

    notified = 0
    for fu in due_followups:
        try:
            # WebSocket push notification
            send_agent_notification(
                schema_name=schema_name,
                agent_id=fu.assigned_to_id,
                event_type="followup_reminder",
                data={
                    "followup_id": fu.pk,
                    "lead_id": fu.lead_id,
                    "lead_name": fu.lead.name,
                    "followup_type": fu.followup_type,
                    "scheduled_at": fu.scheduled_at.isoformat(),
                    "notes": fu.notes,
                },
            )
            fu.reminder_sent = True
            fu.reminder_sent_at = now
            fu.save(update_fields=["reminder_sent", "reminder_sent_at"])
            notified += 1
        except Exception as exc:
            logger.warning(f"[Task] Reminder failed for follow-up {fu.pk}: {exc}")

    logger.info(
        f"[Task] Follow-up reminders sent: {notified} in {schema_name}"
    )
    return {"notified": notified}


# ============================================================
# Data Retention Cleanup
# ============================================================

@shared_task(base=TenantAwareTask, bind=True)
def cleanup_old_leads(self, schema_name: str, retention_days: int):
    """
    Archive (soft-delete) leads older than the plan's data retention period.
    Called monthly for each tenant with their plan's retention_days.
    """
    from datetime import timedelta
    from apps.leads.models import Lead

    cutoff = timezone.now() - timedelta(days=retention_days)
    count = Lead.objects.filter(
        created_at__lt=cutoff,
        is_deleted=False,
    ).update(is_deleted=True)

    logger.info(
        f"[Task] Data retention: archived {count} leads in {schema_name} "
        f"(retention: {retention_days} days)"
    )
    return {"archived": count}


@shared_task(base=PublicSchemaTask, bind=True)
def dispatch_lead_score_recalculation(self):
    """Beat dispatcher: trigger lead score recalculation for every tenant."""
    from apps.core.utils import get_all_tenant_schemas
    schemas = get_all_tenant_schemas()
    for schema_name in schemas:
        recalculate_lead_scores.apply_async(args=[schema_name], queue="default")
    return {"dispatched": len(schemas)}
